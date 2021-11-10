import sys, os, time, datetime, openpyxl
from openpyxl import load_workbook
#from Dymo import Dymoprint as DP
from Test_Lists import csvparse as parse
from PDFcreate import pdfcreate
#Account class directs what each account uses for its documents and user info
class Account:
    def __init__(self, name, username, password, stoolDoc, DLreq, ASTreq, amclreq):
        self.name = name 
        self.username = username
        self.password = password
        self.stool = stoolDoc
        self.blankreq = DLreq
        self.ast = ASTreq
        self.amcl = amclreq


#printing document function, takes string arg of the full filepath of the file
def printdoc(filepath):
    os.startfile(filepath, "print")
    time.sleep(2)

#creates DYMO label that will be printed. Search_for is the user defined text in the dymo .label template file, replace is what we replace it with
def label_create(search_for, replace_with):

    f = open(templateL, 'r')
    filedata = f.read()
    f.close

    newdata = filedata.replace(search_for, replace_with)

    f = open(newL, 'w')
    f.write(newdata)

    f.close()
'''
Right now im not using this, I actually found the xl sheet to be more readable with the regular list format ie ['item1', 'item2', 'etc']
#simple function to assign contents of a list and returns str "item1; item2; item3". Used for the list of user input tests
def cutbracket(listz, target):
    for item in listz:
        target += (item + "; ")
    return target
    
To cut just the brackets off the list use the following function:

def cutbrax(listz, target):
    lstring = str(listz)
    removed = filter(lambda char: char not in "[]", lstring)
    for ch in removed:
        target += char
    return target
    '''
#function that displays items in a dictionary for the user to select, these selected items are appended to a list.
def testappend(dictname, targetlist, targetmod): #emptydictionary, emptylist, list that the label numbers will be added up from

    #display list of tests from the dictionary
    for i in dictname:
        print(i, ":", dictname[i][0])

    repeatwrin = "y"
    repeat = "y"    
    while repeat == "y":
        sel = input("\nCode: ")
        
        if sel in dictname:
            targetlist.append(sel + ": " + dictname[sel][0])
            targetmod.append(int(dictname[sel][1])) #label number from csv sheet is pulled and added to the modifier list
        else:
            print("That is not on the list...try again")
        #0 used as user input for writins, runs a similar loop so custom tests can be plugged in
        if "0: Write in Test" in targetlist:           
            while repeatwrin == "y":
                wrin = input('Please enter the code and testname format:"CODE:TESTNAME": ')
                targetlist.append(wrin + "-w")
                repeatwrin = input("\nAdd another writein? y/n: ").lower()
        
        repeat = input("\nAdd another test? y/n: ").lower()

        if "0: Write in Test" in targetlist: targetlist.remove("0: Write in Test")
#end funcblock***********************************************************************************************


#Directories for each requisition type. Might add to anoter csv later so the user can edit easier.
mydir = os.path.realpath(os.path.join(os.getcwd(), os.path.dirname(__file__)))
DL_Stool = "./Documents to Print/DL_Stool.docx"
P1_Stool = "./Documents to Print/P1_Stool.docx"
L1_Blank = "./Documents to Print/template_L1.pdf"
P1_Blank = "./Documents to Print/template_L1.pdf"
P2_Blank = "./Documents to Print/template_P2.pdf"
L1_AST = 0
P1_AST = 0
P2_AST = 0
P1_Stool = "./Documents to Print/P1_Stool.docx"
AMCLL1_Blank = "./Documents to Print/L1_blank.docx"
AMCLP1_Blank = "./Documents to Print/P1_blank.docx" #temp reqs, amcl not added yet
AMCLNP_Blank = "./Documents to Print/P2_blank.docx"

#initialize classes for each billing type with the indented requisitions to print
L1 = Account("L1", "1485", "lab999", L1_Stool, L1_Blank, AMCLL1_Blank, L1_AST)
P2 = Account("P2", "1485", "lab999", L1_Stool, P2_Blank, AMCLL1_Blank, P2_AST)
P1 = Account("P1", "1248", "P11234", P1_Stool, P1_Blank, AMCLP1_Blank, P1_AST)
NewPrac = Account("New Practice", "1322", "nppassword", L1_Stool, P2_Blank, AMCLNP_Blank, P2_AST)




fulldate = datetime.date.today()
year = fulldate.strftime('%Y')
month = fulldate.strftime('%B')
date = fulldate.strftime('%m-%d-%y')

#create/check for excel sheet folder by date as well as dated folders
if not os.path.exists(f'{mydir}/{year}/{month}'):
    os.makedirs(f"{mydir}/{year}/{month}")

if not os.path.exists(f'{mydir}/{year}/{month}/{date}.xlsx'):
    wb = load_workbook(filename = f"{mydir}/logtemplate.xlsx")
    wb.save(f'{mydir}/{year}/{month}/{date}.xlsx')
  
wb = load_workbook(filename = f'{mydir}/{year}/{month}/{date}.xlsx')
ws = wb.active
#maxrow = ws.max_row 
#end initializationblock***********************************************************************************************



#User enters patient F and L name, dob, doctor. Move quotes around for testing without having to enter every names every time

'''
firstname = input("Patient's First Name:  ")
lastname = input("Patient's Last Name:  ")
dob = input("Patient's date of birth (dd/mm/yyyy):  ")
sex = input("Patient's sex:  ")
doctor = input("Physician's last name:  ")
'''
firstname = "FnameTest"
lastname = "LnameTest"
sex = 'm'
dob = "07/07/1977"
doctor = "TestDoctor"



#combine l and f name
fullname = (lastname + ", " + firstname)

#this is where we create the new dymo label as New.label using the user input patient info
templateL = fr"{mydir}\Dymo\1by3template.label" #template label location
newL = fr"{mydir}\Dymo\New.label" #label that is created location
#label_create("Lastname, Firstname\ndd/mm/yyyy\ndd/mm/1yyy\nProvider Name", ("%s\n%s\n%s\n%s" % (fullname, dob, date, doctor)))

print("\nYou entered: %s, %s: %s with Dr: %s\n" % (firstname, lastname, dob, doctor))

print("Account options: ")
billtypes = {"0": "None Listed", "1": "L1", "2":  "P1", "3":  "Self Pay","4": "P2", "5": "New Practice"}

for i in billtypes:
    print(i,": ",  billtypes[i])

#user inputs number corresponding to the account in the billtypes dictionary

check1 = True
while check1:
    billselect = (input("\nPlease enter the number corresponding to the listing on the top of the order form:  "))
    try:
        billmethod = billtypes[billselect]
        print("\nYou selected %s!" % (billmethod))
        check1 = False
    except:
        print("\nInvalid entry, try again.")

           
#A couple rules to narrow down accounts to one of the four types and check for L1s that have BCBS (which end up being P2 no matter what)
if billmethod == "Self Pay":
    billmethod = "P1"

if billmethod == "L1":
    BCTF = input("\nIs the insurance type BlueCross Blue Sheild? y/n: ")
    BCTF = BCTF.lower()
    if BCTF == "y":
        print("\nThis is supposed to be a P2! Changing Billing Type to P2.")
        billmethod = "P2"
    else:
        print("\nGreat, continuing to requisition printing...")
        
#Print statement for None Listed to instruct user to take chart to billing manager
if billmethod == "None Listed":
    print("\nIllegal Operation!, Please return chart to billing manager to retreive this information!")
    quit = input("\nPress any key to quit...")
    sys.exit()

print("\nPlease make a copy of all docs attached to the front of the chart including insurance verification if given")


#converts user input string and sets account variable to the correct class
accdict = {"L1": L1, "P1": P1, "New Practice": NewPrac, "P2": P2}
if billmethod in accdict:
    account = accdict[billmethod]
    

'''
old, more obvious method that worked just fine but is more annoying to change later
if billMethod == "L1":
    account = L1

if billMethod == "P2":
    account = P2

if billMethod == "P1":
    account = P1

if billMethod == "New Practice":
    account = NewPrac
'''

#Set up each of three test dictinaries by pulling from editable csv files( see /Test_Lists for csvparse.py and data sheets)
labelvals = [] #list that number of labels per test gets appended to

L1_dict = {}
parse.dictmake(f'{mydir}/Test_Lists/L1_list.csv', L1_dict)
L1_input = []
L1_str = ""
print("Please Select Your L1Labs Test by typing in the test code:")
testappend(L1_dict, L1_input, labelvals)



kit_dict = {}
parse.dictmake(f'{mydir}/Test_Lists/kit_list.csv', kit_dict)
kit_input = []
kit_str = ""
print("Please select Take Home kits by typing in the test code: ")
testappend(kit_dict, kit_input, labelvals)


out_dict = {}
parse.dictmake(f'{mydir}/Test_Lists/out_list.csv', out_dict)
out_input = []
out_str = ""
print("Please select Outsourced Tests by typing in the test code: ")
testappend(out_dict, out_input, labelvals)


modifier = sum(labelvals) #take sum of labelvals in list to get total to add on 

    
    
#list of all items being appended into the spreadsheet. Superlist is the row if items being appended
L1_total = len(L1_input)
out_total = len(out_input)
kit_total = len(kit_input)
totaltest = L1_total + out_total + kit_total
               
superlist = [lastname, firstname, dob, sex, doctor, billmethod, cutbrax(L1_input, L1_str), str(out_input), \
                str(kit_input), L1_total, out_total, kit_total, totaltest]

ws.append(superlist)
wb.save(f'{mydir}/{year}/{month}/{date}.xlsx')

#end datasection**************************************************************************
#Print stool req based on account


if "5200" in str(kit_input):
    print("Printing stool requisition to default printer...")
    #printdoc(account.stool)
    
if "1600" in str(kit_input):
    print("Printing AST req")
    #printdoc(account.ast)
    
if out_input:
    print("Printing AMCL requisition to default printer...")
    #printdoc(account.amcl)

if L1_input:
    print("Printing L1 requisition to default printer...")
    # pdfcreate. run requires: lastname, firstname, dob, dos, doctor, L1_list, out_list, kit_list, template_req, file_tag
    pdfcreate.run(lastname, firstname, dob, date, doctor, str(L1_input), str(out_input), str(kit_input), account.blankreq, f'{date}L1')
    #printdoc(account.blankreq)

baselabelnum = 0
extralabelum = 0

#check if user input lists are full and add a label for each req
inputlist = [L1_input, kit_input, out_input]
   
for listin in inputlist:
	if listin:
		baselabelnum += 1
print(baselabelnum, "base amount")

extralabelnum = int(input("Printing %s labels, would you like to add more? How many(type 0 for none)?: " % (baselabelnum +modifier)))
finalnum = (baselabelnum + extralabelnum +modifier)

print(finalnum)
'''
while extralabelnum >0:  
    if extralabelum < 15:
        print("printing dymo patient labels!")
        #DP.dymoprint(finalnum)
    else:
        print("Cant print more than 15 labels at once...for your own safety try again")
        extralabelnum = int(input("Printing %s labels, would you like to add more? How many(type 0 for none)?: " % (baselabelnum)))
        continue
   ''' 

input("Press any key to end")

    



